from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from .models import AddressRow, REQUIRED_COLUMNS


class InputLoadError(ValueError):
    """Raised when an input file cannot be parsed safely."""


def load_rows(path: str | Path) -> list[AddressRow]:
    input_path = Path(path)
    if not input_path.exists():
        raise InputLoadError(f"Input file not found: {input_path}")

    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        records = _load_csv(input_path)
    elif suffix in {".xlsx", ".xlsm"}:
        records = _load_xlsx(input_path)
    else:
        raise InputLoadError("Unsupported input file. Use .csv, .xlsx, or .xlsm")

    return [_record_to_row(index + 2, record) for index, record in enumerate(records)]


def _load_csv(path: Path) -> list[dict[str, object]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        _ensure_columns(reader.fieldnames or [])
        return [dict(row) for row in reader]


def _load_xlsx(path: Path) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised only without optional dep
        raise InputLoadError("XLSX support requires openpyxl. Install project dependencies.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(rows)]
    except StopIteration as exc:
        raise InputLoadError("Input workbook is empty") from exc
    _ensure_columns(headers)
    records: list[dict[str, object]] = []
    for values in rows:
        record = {header: values[index] if index < len(values) else "" for index, header in enumerate(headers)}
        if any(str(value or "").strip() for value in record.values()):
            records.append(record)
    return records


def _ensure_columns(columns: Iterable[str]) -> None:
    normalized = {str(column).strip() for column in columns}
    missing = [column for column in REQUIRED_COLUMNS if column not in normalized]
    if missing:
        raise InputLoadError(f"Missing required columns: {', '.join(missing)}")


def _record_to_row(row_number: int, record: dict[str, object]) -> AddressRow:
    def text(name: str) -> str:
        return str(record.get(name) or "").strip()

    def bool_text(name: str) -> bool:
        return text(name).casefold() in {"1", "true", "yes", "y", "manual_verified", "verified", "확인", "확인됨"}

    return AddressRow(
        row_number=row_number,
        exchange=text("exchange"),
        coin=text("coin"),
        network=text("network"),
        address=text("address"),
        memo_or_tag=text("memo_or_tag"),
        alias=text("alias"),
        owner_name_kr=text("owner_name_kr"),
        status=text("status"),
        memo_verified_absent=bool_text("memo_verified_absent"),
    )
